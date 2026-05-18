"""Tests do helper xlsx_stream."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from app.adapters.sap.parsers.xlsx_stream import iter_xlsx_rows


@pytest.fixture
def simple_xlsx(tmp_path: Path) -> Path:
    p = tmp_path / "simple.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["NAME", "AGE", "CITY"])
    ws.append(["Alice", 30, "RJ"])
    ws.append(["Bob", 25, "SP"])
    wb.save(str(p))
    return p


@pytest.fixture
def multi_sheet_xlsx(tmp_path: Path) -> Path:
    p = tmp_path / "multi.xlsx"
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "First"
    ws1.append(["A"])
    ws1.append([1])
    ws2 = wb.create_sheet("Second")
    ws2.append(["B"])
    ws2.append([2])
    wb.save(str(p))
    return p


def test_yields_dicts_with_headers(simple_xlsx):
    rows = list(iter_xlsx_rows(simple_xlsx))
    assert len(rows) == 2
    assert rows[0] == {"NAME": "Alice", "AGE": 30, "CITY": "RJ"}
    assert rows[1] == {"NAME": "Bob", "AGE": 25, "CITY": "SP"}


def test_sheet_name_selects_correct_sheet(multi_sheet_xlsx):
    """sheet_name parametrizável."""
    rows_first = list(iter_xlsx_rows(multi_sheet_xlsx, sheet_name="First"))
    rows_second = list(iter_xlsx_rows(multi_sheet_xlsx, sheet_name="Second"))
    assert rows_first == [{"A": 1}]
    assert rows_second == [{"B": 2}]


def test_default_sheet_is_first(multi_sheet_xlsx):
    """Sem sheet_name = primeira sheet."""
    rows = list(iter_xlsx_rows(multi_sheet_xlsx))
    assert rows == [{"A": 1}]


def test_expected_headers_passes_when_valid(simple_xlsx):
    """Validação OK quando esperados estão presentes."""
    rows = list(
        iter_xlsx_rows(simple_xlsx, expected_headers=["NAME", "AGE", "CITY"])
    )
    assert len(rows) == 2


def test_expected_headers_raises_when_missing(simple_xlsx):
    """Validação falha quando expected ausente no arquivo."""
    with pytest.raises(ValueError, match="ausentes"):
        list(iter_xlsx_rows(simple_xlsx, expected_headers=["NAME", "MISSING"]))


def test_expected_headers_case_insensitive(simple_xlsx):
    """Validação ignora caixa."""
    rows = list(
        iter_xlsx_rows(simple_xlsx, expected_headers=["name", "age", "city"])
    )
    assert len(rows) == 2


def test_skip_empty_rows_true_by_default(tmp_path):
    """Linhas (None, None, None) são skipadas por default."""
    p = tmp_path / "with_blanks.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["A", "B"])
    ws.append([1, "x"])
    ws.append([None, None])
    ws.append([2, "y"])
    wb.save(str(p))

    rows = list(iter_xlsx_rows(p))
    assert len(rows) == 2
    assert rows[0]["A"] == 1
    assert rows[1]["A"] == 2


def test_empty_file_yields_nothing(tmp_path):
    """XLSX sem rows = generator vazio."""
    p = tmp_path / "empty.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["A"])  # só header
    wb.save(str(p))

    rows = list(iter_xlsx_rows(p))
    assert rows == []
