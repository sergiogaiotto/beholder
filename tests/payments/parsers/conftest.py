"""Fixtures pra tests dos parsers SAP/WF.

Gera in-memory:
  - msrv5_sample_path: TXT cp1252 com formato real do MSRV5 (separadores,
    header repetido, footer, rows válidas e malformadas)
  - wf_analytics_sample_path: XLSX com os 81 headers oficiais + 2 rows
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from app.adapters.sap.parsers.wf_analytics import WF_ANALYTICS_EXPECTED_HEADERS


# Reproduz a estrutura observada em Pré-B §2.2:
# header → ─── → 2 rows válidas → ─── → footer → 1 malformed → 1 data inválida
MSRV5_SAMPLE_LINES: list[str] = [
    "----------------------------------------------------------------------",
    "| Data doc.  | N° de docu | Item | Serviço | Qtd.   | Preço bruto | Texto breve         |",
    "----------------------------------------------------------------------",
    "| 06.09.2022 | 5700012782 | 7913 | 9000507 | 0,000  |  2,71       | SERV CONFECCAO MAT  |",
    "| 15.05.2023 | 5700099999 | 1    | 9001000 | 1,500  |  150,75     | OUTRO SERVICO       |",
    "----------------------------------------------------------------------",
    "Estat: 2 transferidos",
    "| MALFORMED | 5700099 |",
    "| 31.02.2024 | X | 1 | S | 0,000 | 1,00 | data invalida |",
    "",
    "| 01.01.2024 | 5700111111 | 5    | 9002000 | 0,000  |  99,99      | TERCEIRA VALIDA     |",
]


@pytest.fixture
def msrv5_sample_path(tmp_path: Path) -> Path:
    """TXT MSRV5 com 3 rows válidas + ruído real."""
    p = tmp_path / "msrv5_sample.txt"
    content = "\n".join(MSRV5_SAMPLE_LINES) + "\n"
    p.write_text(content, encoding="cp1252")
    return p


def _build_wf_row(**values) -> list:
    """Constrói linha do XLSX WF a partir de kwargs (header → valor).

    Posiciona valores nas colunas certas; resto fica None.
    """
    headers = list(WF_ANALYTICS_EXPECTED_HEADERS)
    row: list = [None] * len(headers)
    for k, v in values.items():
        row[headers.index(k)] = v
    return row


@pytest.fixture
def wf_analytics_sample_path(tmp_path: Path) -> Path:
    """XLSX WF com os 81 headers oficiais + 2 rows de sample."""
    p = tmp_path / "wf_sample.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Analitico_Empreiteiras_WF1_WF2_"
    ws.append(list(WF_ANALYTICS_EXPECTED_HEADERS))

    from datetime import datetime
    ws.append(_build_wf_row(
        SISTEMA="WF1", OS="OS-001", EMPREITEIRA="ABILITY",
        CONTRATO_NUM="4600012345", ITEM_NUM=10,
        UF="RJ", CIDADE="Rio de Janeiro",
        STATUS_OS="EXECUTADO", NIVEL_GERENCIAL="Em Pagamento",
        MALOGRO="NAO",
        DATA_PEDIDO=datetime(2025, 6, 1),
        VALOR_TOTAL_FINAL=1500.00, VALOR_UNITARIO=125.50,
        MATERIAL_SERVICO_NUM="9000507",
        MES_MEDICAO="2025/06",
    ))
    ws.append(_build_wf_row(
        SISTEMA="WF2", OS="OS-002", EMPREITEIRA="BETA",
        STATUS_OS="EM EXECUÇÃO",
        DATA_PEDIDO=datetime(2025, 7, 15),
    ))

    wb.save(str(p))
    return p


@pytest.fixture
def wf_extra_cols_xlsx(tmp_path: Path) -> Path:
    """XLSX WF com 1 col EXTRA além das 81 oficiais — schema forward-compat."""
    p = tmp_path / "wf_extra.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Analitico_Empreiteiras_WF1_WF2_"
    headers = list(WF_ANALYTICS_EXPECTED_HEADERS) + ["NOVO_CAMPO_FUTURO"]
    ws.append(headers)
    ws.append(_build_wf_row(SISTEMA="WF1", OS="OS-X") + ["extra"])
    wb.save(str(p))
    return p


@pytest.fixture
def wf_missing_cols_xlsx(tmp_path: Path) -> Path:
    """XLSX WF FALTANDO uma col oficial — deve falhar validação."""
    p = tmp_path / "wf_missing.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Analitico_Empreiteiras_WF1_WF2_"
    headers = [h for h in WF_ANALYTICS_EXPECTED_HEADERS if h != "SISTEMA"]
    ws.append(headers)
    # Row com pelo menos 1 valor não-None pra não ser skipada por skip_empty_rows
    sample = [None] * len(headers)
    sample[headers.index("OS")] = "OS-NOSCHEMA"
    ws.append(sample)
    wb.save(str(p))
    return p
