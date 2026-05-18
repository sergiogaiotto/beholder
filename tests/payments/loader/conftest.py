"""Fixtures pra tests integrados do loader.

Reusa o pattern do tests/payments/repositories/conftest:
  - init schema (session autouse)
  - TRUNCATE entre tests
  - test_user_id function-scope
Cria também fixtures de arquivo (XLSX/TXT temporários) usando os
mesmos builders do parsers conftest.
"""

from __future__ import annotations

import uuid
from collections.abc import AsyncGenerator
from datetime import datetime
from pathlib import Path
from uuid import UUID

import pytest
from openpyxl import Workbook

from app.adapters.db.postgres import connect
from app.adapters.db.postgres_payments import connect_payments, init_payments_schema
from app.adapters.sap.parsers.wf_analytics import WF_ANALYTICS_EXPECTED_HEADERS


@pytest.fixture(scope="session", autouse=True)
def _init_payments_for_loader(event_loop):
    event_loop.run_until_complete(init_payments_schema())


@pytest.fixture(autouse=True)
async def _reset_payments_per_test():
    """TRUNCATE entre tests + limpa catálogos de teste — espelha repositories conftest."""
    async with connect_payments() as c:
        await c.execute(
            """
            TRUNCATE
                payments.reconciliation_finding,
                payments.analytic_finding,
                payments.reconciliation_run,
                payments.extraction_job,
                payments.contract_clause,
                payments.lpu_item,
                payments.contract_version,
                payments.contract_master,
                payments.supplier_bridge,
                payments.purchase_order_item,
                payments.purchase_order_header,
                payments.service_package,
                payments.purchase_order_gc,
                payments.cost_center_account,
                payments.wf_payment,
                payments.ingestion_run
            RESTART IDENTITY CASCADE
            """
        )
    yield


@pytest.fixture
async def test_user_id() -> AsyncGenerator[UUID, None]:
    user_id = uuid.uuid4()
    username = f"loader_test_{uuid.uuid4().hex[:8]}"
    async with connect() as db:
        await db.execute(
            """
            INSERT INTO users (id, username, hashed_password, salt, is_active)
            VALUES ($1::uuid, $2, 'x', 'x', TRUE)
            """,
            str(user_id), username,
        )
    yield user_id


# ---------- File fixtures ----------


def _build_wf_row(**values) -> list:
    headers = list(WF_ANALYTICS_EXPECTED_HEADERS)
    row: list = [None] * len(headers)
    for k, v in values.items():
        row[headers.index(k)] = v
    return row


@pytest.fixture
def supplier_bridge_xlsx(tmp_path: Path) -> Path:
    """XLSX com schema do Contratos-Empreteiras (sheet 'Empreiteiras')."""
    p = tmp_path / "supplier.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Empreiteiras"
    ws.append([
        "CATEGORIA", "EMPREITEIRA", "CONTRATO_NUM", "REF WS",
        "NUMERO_FORNECEDOR SAP", "CNPJ",
    ])
    ws.append(["OBRAS", "ABILITY", "4600012345", "WS-001", "100200", "12345678000199"])
    ws.append(["OBRAS", "BETA", "4600099999", "WS-002", "200300", "98765432000188"])
    wb.save(str(p))
    return p


@pytest.fixture
def cost_center_xlsx(tmp_path: Path) -> Path:
    """XLSX com sheet 'CC + CONTA'."""
    p = tmp_path / "cost_center.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "CC + CONTA"
    ws.append(["CENTRO_DE_CUSTO", "CONTA_RAZAO"])
    ws.append(["CC-1", "6010101"])
    ws.append(["CC-1", "6010102"])
    ws.append(["CC-2", "6010101"])
    wb.save(str(p))
    return p


@pytest.fixture
def wf_payment_xlsx(tmp_path: Path) -> Path:
    """XLSX WF analítico com 81 headers + 3 rows."""
    p = tmp_path / "wf_payment.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Analitico_Empreiteiras_WF1_WF2_"
    ws.append(list(WF_ANALYTICS_EXPECTED_HEADERS))
    ws.append(_build_wf_row(
        SISTEMA="WF1", OS="OS-1", EMPREITEIRA="ABILITY",
        DATA_PEDIDO=datetime(2025, 6, 1),
        VALOR_TOTAL_FINAL=1500.00,
        STATUS_OS="EXECUTADO", NIVEL_GERENCIAL="Em Pagamento", MALOGRO="NAO",
    ))
    ws.append(_build_wf_row(
        SISTEMA="WF2", OS="OS-2", EMPREITEIRA="BETA",
        DATA_PEDIDO=datetime(2025, 7, 15),
        STATUS_OS="EXECUTADO", NIVEL_GERENCIAL="Em Pagamento", MALOGRO="NAO",
    ))
    ws.append(_build_wf_row(
        SISTEMA="WF1", OS="OS-3", EMPREITEIRA="ABILITY",
        DATA_PEDIDO=datetime(2026, 1, 15),
        STATUS_OS="CANCELADO",  # fora do filtro universal
    ))
    wb.save(str(p))
    return p


@pytest.fixture
def msrv5_txt(tmp_path: Path) -> Path:
    """TXT MSRV5 cp1252 com 4 rows válidas + ruído (2 anos pra cobrir partições)."""
    p = tmp_path / "msrv5.txt"
    lines = [
        "----------------------------------------------------------------------",
        "| Data doc.  | N° de docu | Item | Serviço | Qtd.   | Preço bruto | Texto breve         |",
        "----------------------------------------------------------------------",
        "| 06.09.2022 | 5700012782 | 7913 | 9000507 | 0,000  |  2,71       | SERV CONFECCAO MAT  |",
        "| 15.05.2023 | 5700099999 | 1    | 9001000 | 1,500  |  150,75     | OUTRO SERVICO       |",
        "| 01.01.2024 | 5700111111 | 5    | 9002000 | 0,000  |  99,99      | TERCEIRA VALIDA     |",
        "| 10.03.2025 | 5700222222 | 2    | 9003000 | 2,000  |  500,00     | QUARTA VALIDA       |",
        "----------------------------------------------------------------------",
    ]
    p.write_text("\n".join(lines) + "\n", encoding="cp1252")
    return p
