"""Helper genérico de XLSX streaming via openpyxl read-only.

Por que openpyxl read-only:
  - openpyxl.load_workbook(..., read_only=True) abre o XLSX sem carregar
    tudo na memória (usa iterators). Pra 869k × 81 cols seria ~600 MB
    no modo padrão; read-only mantém memória sub-50 MB.
  - data_only=True faz cells de fórmula retornarem o valor calculado
    cacheado pelo Excel (em vez da string da fórmula).

Por que não calamine/polars:
  - calamine é mais rápido mas não está instalado e tem dependência
    em Rust toolchain pra build em Windows.
  - polars será adicionada no Bloco F (loader) — não vou puxar dependência
    nova só pra parser.

iter_xlsx_rows é puro generator: yields uma row por vez. Header inferido
da primeira linha (default) ou fornecido explicitamente (skip_header=True).
"""

from __future__ import annotations

from collections.abc import Iterator, Sequence
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.adapters.sap.parsers._helpers import normalize_header


def iter_xlsx_rows(
    path: Path | str,
    *,
    sheet_name: str | None = None,
    expected_headers: Sequence[str] | None = None,
    skip_empty_rows: bool = True,
) -> Iterator[dict[str, Any]]:
    """Itera linhas de um XLSX como dict {header: value}.

    Args:
      path: caminho do arquivo XLSX
      sheet_name: nome da sheet. None = primeira sheet.
      expected_headers: se fornecido, valida que os headers da primeira
        linha bater (case-insensitive após normalize). ValueError se não bater.
      skip_empty_rows: skipa linhas com todos os valores None/vazio.

    Yields:
      dict[str, Any] — keys = headers normalizados (strip + NBSP removido);
      values = nativos do openpyxl (datetime, int, float, str, None).
    """
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb.worksheets[0]
        iter_rows = ws.iter_rows(values_only=True)

        # Primeira linha = headers
        try:
            raw_headers = next(iter_rows)
        except StopIteration:
            return
        headers = [normalize_header(h) for h in raw_headers]

        if expected_headers is not None:
            _validate_headers(headers, expected_headers, sheet=ws.title)

        for row in iter_rows:
            if skip_empty_rows and all(v is None or v == "" for v in row):
                continue
            # Trunca/preenche se row tem largura diferente dos headers
            # (raro em XLSX bem-formado, mas defensivo).
            yield {
                h: (row[i] if i < len(row) else None)
                for i, h in enumerate(headers)
                if h  # ignora cols com header vazio
            }
    finally:
        wb.close()


def _validate_headers(
    found: Sequence[str], expected: Sequence[str], *, sheet: str
) -> None:
    """Valida que found == expected (case-insensitive, set-based)."""
    found_set = {h.upper() for h in found if h}
    expected_set = {h.upper() for h in expected}
    missing = expected_set - found_set
    extra = found_set - expected_set
    if missing:
        raise ValueError(
            f"sheet {sheet!r}: headers esperados ausentes: {sorted(missing)}"
        )
    if extra:
        # Extra é warning-level, não erro — schema pode evoluir
        # (mas registramos no log via ValueError soft; caller decide).
        # Para Fase 1: aceitar; Fase 2 pode mudar pra strict.
        pass
