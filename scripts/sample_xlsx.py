"""Deep sample de uma sheet XLSX: todos os headers + amostras + valores únicos
em colunas seletas. Usado na Pré-B para sondar Analítico WF e Casos Selecionados.

Uso:
    .venv\\Scripts\\python.exe scripts\\sample_xlsx.py <xlsx_path> <sheet_name> [--samples N]
"""
from __future__ import annotations

import argparse
import json
import sys
from collections import Counter
from pathlib import Path

import openpyxl  # type: ignore[import-untyped]

SAMPLE_DEFAULT = 100


def sample_sheet(path: Path, sheet_name: str, n_samples: int) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return {"error": f"sheet '{sheet_name}' not found", "sheets": wb.sheetnames}
        ws = wb[sheet_name]
        out: dict = {
            "sheet": sheet_name,
            "declared_rows": ws.max_row or 0,
            "declared_cols": ws.max_column or 0,
            "headers": [],
            "samples": [],
            "non_null_count_per_col": [],
            "unique_count_per_col": [],
        }
        iter_rows = ws.iter_rows(values_only=True)
        try:
            headers = next(iter_rows)
        except StopIteration:
            return out
        out["headers"] = [str(c) if c is not None else "" for c in headers]
        ncols = len(headers)

        non_null = [0] * ncols
        unique_counters: list[set] = [set() for _ in range(ncols)]
        # Limit unique tracking to avoid memory blow-up
        unique_cap = 1000

        # Sample first n_samples rows
        for i, row in enumerate(iter_rows):
            row_values = list(row) if row is not None else []
            # pad row to ncols
            while len(row_values) < ncols:
                row_values.append(None)
            for j, v in enumerate(row_values[:ncols]):
                if v is not None and v != "":
                    non_null[j] += 1
                    if len(unique_counters[j]) < unique_cap:
                        unique_counters[j].add(str(v))
            if i < n_samples:
                out["samples"].append([str(v) if v is not None else "" for v in row_values[:ncols]])
            if i % 100_000 == 0 and i > 0:
                print(f"  ... {i} rows scanned", file=sys.stderr, flush=True)

        out["non_null_count_per_col"] = non_null
        out["unique_count_per_col"] = [
            (len(u) if len(u) < unique_cap else f"≥{unique_cap}") for u in unique_counters
        ]
        # Top values for small-cardinality columns (likely categorical)
        out["categorical_preview"] = {}
        for j, hdr in enumerate(out["headers"]):
            if isinstance(out["unique_count_per_col"][j], int) and out["unique_count_per_col"][j] <= 50:
                out["categorical_preview"][hdr] = sorted(unique_counters[j])[:50]
        return out
    finally:
        wb.close()


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("xlsx_path", type=Path)
    p.add_argument("sheet_name", type=str)
    p.add_argument("--samples", type=int, default=SAMPLE_DEFAULT)
    p.add_argument("--out", type=Path, default=None)
    args = p.parse_args()

    if not args.xlsx_path.exists():
        print(f"ERRO: {args.xlsx_path} não existe", file=sys.stderr)
        return 1

    print(f"Sondando {args.xlsx_path.name} / sheet '{args.sheet_name}' ...", file=sys.stderr)
    result = sample_sheet(args.xlsx_path, args.sheet_name, args.samples)
    out_json = json.dumps(result, ensure_ascii=False, indent=2, default=str)
    if args.out:
        args.out.write_text(out_json, encoding="utf-8")
        print(f"\nGravado em {args.out}", file=sys.stderr)
    else:
        print(out_json)
    return 0


if __name__ == "__main__":
    sys.exit(main())
